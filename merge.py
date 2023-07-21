# from openpyxl import load_workbook
# import os

# # Define the directory where your XLSX files are located
# directory = './folder_all'

# # Create a new workbook to store the merged data
# merged_workbook = load_workbook(filename='template.xlsx')

# # Iterate over each file in the directory
# for filename in os.listdir(directory):
#     if filename.endswith('.xlsx'):
#         file_path = os.path.join(directory, filename)
#         # Load the source workbook
#         source_workbook = load_workbook(file_path, read_only=True)
#         # Iterate over each sheet in the source workbook
#         for sheet_name in source_workbook.sheetnames:
#             # Get the source sheet
#             source_sheet = source_workbook[sheet_name]
#             # Create a new sheet in the merged workbook with the same name
#             merged_sheet = merged_workbook.create_sheet(title=sheet_name)
#             # Iterate over each row in the source sheet
#             for row in source_sheet.iter_rows(values_only=True):
#                 # Append the row to the merged sheet
#                 merged_sheet.append(row)

# # Remove the default sheet created in the merged workbook
# merged_workbook.remove(merged_workbook['ListOfTenders'])

# # Save the merged workbook to a new XLSX file
# merged_workbook.save('merged.xlsx')



from openpyxl import load_workbook
import os

# Define the directory where your XLSX files are located
directory = './folder_all'

# Create a new workbook to store the merged data
merged_workbook = load_workbook(filename='merged.xlsx')

# Create a new sheet in the merged workbook
merged_sheet = merged_workbook.active

# Iterate over each file in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)
        # Load the source workbook
        source_workbook = load_workbook(file_path, read_only=True)
        # Iterate over each sheet in the source workbook
        for sheet_name in source_workbook.sheetnames:
            # Get the source sheet
            source_sheet = source_workbook[sheet_name]
            # Iterate over each row in the source sheet
            for row in source_sheet.iter_rows(values_only=True):
                # Append the row to the merged sheet
                merged_sheet.append(row)

# Save the merged workbook to a new XLSX file
merged_workbook.save('merged.xlsx')
